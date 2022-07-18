from bs4 import BeautifulSoup
import requests
import openpyxl

excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = 'Goodreads Books'
print(excel.sheetnames)
sheet.append(['Rating and review 1'])

url1 = "https://www.goodreads.com/book/show/52845775-minor-feelings"
r = requests.get(url1)
source = r.content
soup = BeautifulSoup(source, 'html.parser')
rating = soup.find('div', id = 'metacol', class_ = 'last col').find('div', id = 'bookMeta', class_ = 'uitext stacked')
# print(rating)
rate = soup.find('span', itemprop = 'ratingValue').get_text(strip = True)
print(rate)
rating_s = soup.find('a', class_ = 'gr-hyperlink')
rate_s = soup.find('meta', itemprop = 'ratingCount').get_text(strip = True)
print(rate_s)
review = soup.find('a', class_ = 'gr-hyperlink')
rev = soup.find('meta', itemprop = 'reviewCount').get_text(strip = True)
print(rev)
sheet.append([rate, rate_s, rev])
excel.save('Books.xlsx')