from bs4 import BeautifulSoup
import requests
import openpyxl

excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = 'Goodreads Books'
print(excel.sheetnames)
sheet.append(['Book Title', 'Ratings', 'Reviews'])

url = "https://www.goodreads.com/list/show/147668.E_J_Koh_s_Books_to_Celebrate_Asian_American_Fiction_Non_Fiction_Memoir_Graphic_Novel_and_Poetry"
r = requests.get(url)
source = r.content
soup = BeautifulSoup(source, 'html.parser')
books = soup.find('table', class_= "tableList js-dataTooltip").find_all('tr')
"""print(len(books))"""
for book in books:
    name = book.find("a", class_ = "bookTitle").get_text(strip=True)
    print(name)
    #ratingsss and reviewss
print('*****1******')
url1 = "https://www.goodreads.com/book/show/52845775-minor-feelings"
r = requests.get(url1)
source = r.content
soup = BeautifulSoup(source, 'html.parser')
rating = soup.find('div', id = 'metacol', class_ = 'last col').find('div', id = 'bookMeta', class_ = 'uitext stacked')
rate = soup.find('span', itemprop = 'ratingValue').get_text(strip = True)
print(rate)
rating_s = soup.find('a', class_ = 'gr-hyperlink')
rate_s = soup.find('meta', itemprop = 'ratingCount').get_text(strip = True)
print(rate_s)
review = soup.find('a', class_ = 'gr-hyperlink')
rev = soup.find('meta', itemprop = 'reviewCount').get_text(strip = True)
print(rev)

#rating2
print('******2******')
url2 = "https://www.goodreads.com/book/show/54578808-why-do-you-look-at-me-and-see-a-girl"
r2 = requests.get(url2)
source2 = r2.content
soup2 = BeautifulSoup(source2, 'html.parser')
rating2 = soup2.find('div', id = 'metacol', class_ = 'last col').find('div', id = 'bookMeta', class_ = 'uitext stacked')
rate2 = soup2.find('span', itemprop = 'ratingValue').get_text(strip = True)
print(rate2)
rating_s2 = soup2.find('a', class_ = 'gr-hyperlink')
rate_s2 = soup2.find('meta', itemprop = 'ratingCount').get_text(strip = True)
print(rate_s2)

review2 = soup2.find('a', class_ = 'gr-hyperlink')
rev2 = soup2.find('meta', itemprop = 'reviewCount').get_text(strip = True)
print(rev2)

#rating3
print('******3*******')
url3 = 'https://www.goodreads.com/book/show/53029244-obit'
r3 = requests.get(url3)
source3 = r3.content
soup3 = BeautifulSoup(source3, 'html.parser')
rating3 = soup3.find('div', id = 'metacol', class_ = 'last col').find('div', id = 'bookMeta', class_ = 'uitext stacked')
rate3 = soup3.find('span', itemprop = 'ratingValue').get_text(strip = True)
print(rate3)
rating_s3 = soup3.find('a', class_ = 'gr-hyperlink')
rate_s3 = soup3.find('meta', itemprop = 'ratingCount').get_text(strip = True)
print(rate_s3)

review3 = soup3.find('a', class_ = 'gr-hyperlink')
rev3 = soup3.find('meta', itemprop = 'reviewCount').get_text(strip = True)
print(rev3)

#rating4
print('******4*******')
url4 = 'https://www.goodreads.com/book/show/54578818-fuse'
r4 = requests.get(url4)
source4 = r4.content
soup4 = BeautifulSoup(source4, 'html.parser')
rating4 = soup4.find('div', id = 'metacol', class_ = 'last col').find('div', id = 'bookMeta', class_ = 'uitext stacked')
rate4 = soup4.find('span', itemprop = 'ratingValue').get_text(strip = True)
print(rate4)
rating_s34= soup4.find('a', class_ = 'gr-hyperlink')
rate_s4 = soup4.find('meta', itemprop = 'ratingCount').get_text(strip = True)
print(rate_s4)

review4 = soup4.find('a', class_ = 'gr-hyperlink')
rev4 = soup4.find('meta', itemprop = 'reviewCount').get_text(strip = True)
print(rev4)

#rating5
print('******5*******')
url5 = 'https://www.goodreads.com/book/show/40030311-almost-american-girl'
r5 = requests.get(url5)
source5 = r5.content
soup5 = BeautifulSoup(source5, 'html.parser')
rating5 = soup5.find('div', id = 'metacol', class_ = 'last col').find('div', id = 'bookMeta', class_ = 'uitext stacked')
rate5 = soup5.find('span', itemprop = 'ratingValue').get_text(strip = True)
print(rate5)
rating_s5= soup5.find('a', class_ = 'gr-hyperlink')
rate_s5 = soup5.find('meta', itemprop = 'ratingCount').get_text(strip = True)
print(rate_s5)

review5 = soup5.find('a', class_ = 'gr-hyperlink')
rev5 = soup5.find('meta', itemprop = 'reviewCount').get_text(strip = True)
print(rev5)

#rating6
print('******6*******')
url6 = 'https://www.goodreads.com/book/show/54578811-letters-from-johnny'
r6 = requests.get(url6)
source6 = r6.content
soup6 = BeautifulSoup(source6, 'html.parser')
rating6 = soup6.find('div', id = 'metacol', class_ = 'last col').find('div', id = 'bookMeta', class_ = 'uitext stacked')
rate6 = soup6.find('span', itemprop = 'ratingValue').get_text(strip = True)
print(rate6)
rating_s6= soup6.find('a', class_ = 'gr-hyperlink')
rate_s6 = soup6.find('meta', itemprop = 'ratingCount').get_text(strip = True)
print(rate_s6)

review6 = soup6.find('a', class_ = 'gr-hyperlink')
rev6 = soup6.find('meta', itemprop = 'reviewCount').get_text(strip = True)
print(rev6)

#ratings7
print('******7*******')
url7 = 'https://www.goodreads.com/book/show/46144864-a-map-is-only-one-story'
r7 = requests.get(url7)
source7 = r7.content
soup7 = BeautifulSoup(source7, 'html.parser')
rating7 = soup7.find('div', id = 'metacol', class_ = 'last col').find('div', id = 'bookMeta', class_ = 'uitext stacked')
rate7 = soup7.find('span', itemprop = 'ratingValue').get_text(strip = True)
print(rate7)
rating_s7= soup7.find('a', class_ = 'gr-hyperlink')
rate_s7 = soup7.find('meta', itemprop = 'ratingCount').get_text(strip = True)
print(rate_s7)

review7 = soup7.find('a', class_ = 'gr-hyperlink')
rev7 = soup7.find('meta', itemprop = 'reviewCount').get_text(strip = True)
print(rev7)

#ratings8
print('******8*******')
url8 = 'https://www.goodreads.com/book/show/54578815-the-language-we-were-never-taught-to-speak'
r8 = requests.get(url8)
source8 = r8.content
soup8 = BeautifulSoup(source8, 'html.parser')
rating8 = soup8.find('div', id = 'metacol', class_ = 'last col').find('div', id = 'bookMeta', class_ = 'uitext stacked')
rate8 = soup8.find('span', itemprop = 'ratingValue').get_text(strip = True)
print(rate8)
rating_s8= soup8.find('a', class_ = 'gr-hyperlink')
rate_s8 = soup8.find('meta', itemprop = 'ratingCount').get_text(strip = True)
print(rate_s8)

review8 = soup8.find('a', class_ = 'gr-hyperlink')
rev8 = soup8.find('meta', itemprop = 'reviewCount').get_text(strip = True)
print(rev8)

#ratings9
print('******9*******')
url9 = 'https://www.goodreads.com/book/show/45010953-a-nail-the-evening-hangs-on'
r9 = requests.get(url9)
source9 = r9.content
soup9 = BeautifulSoup(source9, 'html.parser')
rating9 = soup9.find('div', id = 'metacol', class_ = 'last col').find('div', id = 'bookMeta', class_ = 'uitext stacked')
rate9 = soup9.find('span', itemprop = 'ratingValue').get_text(strip = True)
print(rate9)
rating_s9= soup9.find('a', class_ = 'gr-hyperlink')
rate_s9 = soup9.find('meta', itemprop = 'ratingCount').get_text(strip = True)
print(rate_s9)

review9 = soup9.find('a', class_ = 'gr-hyperlink')
rev9 = soup9.find('meta', itemprop = 'reviewCount').get_text(strip = True)
print(rev9)

#ratings10
print('******10*******')
url10 = 'https://www.goodreads.com/book/show/52568667-tales-from-the-bottom-of-my-sole'
r10 = requests.get(url10)
source10 = r10.content
soup10 = BeautifulSoup(source10, 'html.parser')
rating10 = soup10.find('div', id = 'metacol', class_ = 'last col').find('div', id = 'bookMeta', class_ = 'uitext stacked')
rate10 = soup10.find('span', itemprop = 'ratingValue').get_text(strip = True)
print(rate10)
rating_s10= soup10.find('a', class_ = 'gr-hyperlink')
rate_s10 = soup10.find('meta', itemprop = 'ratingCount').get_text(strip = True)
print(rate_s10)

review10 = soup10.find('a', class_ = 'gr-hyperlink')
rev10 = soup10.find('meta', itemprop = 'reviewCount').get_text(strip = True)
print(rev10)