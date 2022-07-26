from bs4 import BeautifulSoup
import requests
import openpyxl
excel = openpyxl.Workbook()
# print(excel.sheetnames)
sheet = excel.active
sheet.title = 'IMDB Movies'
print(excel.sheetnames)
sheet.append(['Movie Title', 'Rating'])
url1 = "https://www.imdb.com/search/title/?genres=family&sort=user_rating,desc&title_type=feature&num_votes=25000,&pf_rd_m=A2FGELUUNOQJNL&pf_rd_p=5aab685f-35eb-40f3-95f7-c53f09d542c3&pf_rd_r=7G3C5TH0YYSDWWPDQ3YB&pf_rd_s=right-6&pf_rd_t=15506&pf_rd_i=top&ref_=chttp_gnr_8"
r = requests.get(url1)
source = r.content
soup = BeautifulSoup(source, 'html.parser')
Item = soup.find('div', id = 'main').find_all('div', class_ = 'lister-item-content')
#print(len(Title1))
for items in Item:
    name = items.find('h3', class_ = 'lister-item-header').get_text(strip = True)
    print(name)
    rating = items.find('div', class_='ratings-bar').strong.get_text(strip = True)
    print(rating)
    sheet.append([name, rating])
excel.save('Movies.xlsx')
    
    